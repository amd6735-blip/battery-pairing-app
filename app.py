import streamlit as st
import pandas as pd
from itertools import combinations
import plotly.express as px
import numpy as np
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import uuid

FIX_COLOR_MAP = {"Unmatched": "#ffb3b3"}
st.set_page_config(page_title="Dynapack Cell Regroup System", layout="wide", page_icon="🔋")

st.markdown("""
<style>
button[key="run_analysis_btn"] {
    background-color: #072280 !important;
    color: white !important;
    font-size: 20px !important;
    padding: 18px 60px !important;
    border-radius: 12px !important;
    border: none !important;
    width: 100% !important;
    font-weight: 600;
}
button[key="run_analysis_btn"]:hover {
    background-color: #051a60 !important;
    transform: scale(1.02);
    transition: 0.2s;
}
button[key="mock_data_btn"] {
    font-size: 16px;
    padding: 12px;
    width: 100%;
}
</style>
""", unsafe_allow_html=True)

init_keys = [
    "login", "need_reset", "uploader_key", "upload_raw_data", "mock_data_df",
    "final_group_df", "remain_unmatch", "group_summary_df", "batch_report_list",
    "fig_main", "fig_scatter", "total_cells", "matched_unique", "unmatch_total",
    "total_pack_num", "yield_rate"
]
for k in init_keys:
    if k not in st.session_state:
        if k == "login":
            st.session_state[k] = False
        elif k == "need_reset":
            st.session_state[k] = False
        elif k == "uploader_key":
            st.session_state[k] = str(uuid.uuid4())
        elif k in ["upload_raw_data","mock_data_df","final_group_df","remain_unmatch","group_summary_df"]:
            st.session_state[k] = pd.DataFrame()
        else:
            st.session_state[k] = None

if not st.session_state.login:
    st.markdown("""
    <style>
    .login-container {max-width:860px;margin:50px auto;padding:60px 40px;text-align:center;background:linear-gradient(130deg,#fff,#eff5ff);border-radius:24px;box-shadow:0 12px 45px rgba(10,35,100,0.13);}
    .welcome-title {font-size:32px;font-weight:700;color:#072280;margin:32px 0;letter-spacing:0.7px;}
    .stTextInput {max-width:480px;margin:0 auto 30px;}
    .stButton>button {background:#072280;color:white;font-size:17px;padding:13px 45px;border-radius:10px;border:none;}
    </style>
    """, unsafe_allow_html=True)
    with st.container():
        st.markdown('<div class="login-container">', unsafe_allow_html=True)
        try:
            st.image("dynapack_logo.png", width=680)
        except:
            st.markdown("# 🔋 DYNAPACK", unsafe_allow_html=True)
        st.markdown('<div class="welcome-title">Welcome to Dynapack Non-matchable Cell regrouping System </div>', unsafe_allow_html=True)
        pwd = st.text_input("請輸入系統密碼", type="password")
        if st.button("登入系統"):
            if pwd == "3211":
                st.session_state.login = True
                st.rerun()
            else:
                st.error("密碼錯誤")
        st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

def trigger_reset():
    st.session_state.need_reset = True
col_title, col_reset = st.columns([9,1])
with col_title:
    st.title("🔋 Dynapack Non-matchable Cell regrouping System")
with col_reset:
    st.button("🔄 重置所有資料", on_click=trigger_reset, type="secondary")
if st.session_state.need_reset:
    clear_keys = [
        "upload_raw_data","mock_data_df","final_group_df","remain_unmatch","group_summary_df",
        "batch_report_list","fig_main","fig_scatter","total_cells","matched_unique",
        "unmatch_total","total_pack_num","yield_rate"
    ]
    for k in clear_keys:
        if k in ["upload_raw_data","mock_data_df","final_group_df","remain_unmatch","group_summary_df"]:
            st.session_state[k] = pd.DataFrame()
        else:
            st.session_state[k] = None
    st.session_state.uploader_key = str(uuid.uuid4())
    st.session_state.need_reset = False
    st.rerun()

with st.sidebar:
    st.header("⚙️ 分析參數 (Configuration)")
    S = st.number_input("串聯數 (S)", min_value=1, max_value=24, value=14, key="S_input")
    P = st.number_input("並聯數 (P)", min_value=1, max_value=6, value=6, key="P_input")
    base_target = S * P
    base_mv = st.select_slider("Max Delta V 基礎門檻(mV)", options=[1,2,3,4,5,6,7,8,9,10], value=3)
    base_threshold = base_mv / 1000

    st.markdown("#### 🧩 Buffer電芯容差設定")
    enable_buffer = st.checkbox("開啟自放電/設備誤差Buffer容差", value=False)
    buffer_cell = st.select_slider("Buffer備用電芯數", options=[0,1,2,3], value=0)
    if enable_buffer:
        win_len = base_target + buffer_cell
        st.caption(f"已開啟Buffer，單組配組電芯數：{win_len}（基礎{S}×{P} + Buffer{buffer_cell}）")
    else:
        win_len = base_target
        st.caption(f"未開啟Buffer，單組配組電芯數：{win_len}（基礎{S}×{P}，無額外電芯）")

    st.markdown("---")
    st.header("🔬 演算法驗證模式")
    enable_s2 = st.checkbox("Sn Ratio (啟用第二階段混批組配)", value=True)

def load_and_fix_data(uploaded_files):
    all_data = []
    batch_keys = ['batch','批次','批號','batch_no','batchid']
    ocv_keys = ['ocv','電壓','voltage','v','電池電壓']
    if uploaded_files:
        for f in uploaded_files:
            try:
                if f.name.endswith((".xlsx",".xls")):
                    df_raw = pd.read_excel(f)
                else:
                    df_raw = pd.read_csv(f, encoding_errors="replace")
                df_raw.columns = df_raw.columns.str.strip().str.lower()
                b_col, o_col = None, None
                for c in df_raw.columns:
                    if not b_col and any(k in c for k in batch_keys):
                        b_col = c
                    if not o_col and any(k in c for k in ocv_keys):
                        o_col = c
                if not b_col or not o_col:
                    st.error(f"{f.name} 缺少批次/OCV欄位，跳過")
                    continue
                temp = df_raw[[b_col, o_col]].copy()
                temp.columns = ["Batch","OCV"]
                def clean(v):
                    try: return float(v)
                    except: return np.nan
                temp["OCV"] = temp["OCV"].apply(clean)
                temp = temp.dropna(subset=["Batch","OCV"])
                temp = temp[temp["OCV"]>0]
                if len(temp):
                    all_data.append(temp)
            except Exception as e:
                st.error(f"讀取失敗：{e}")
    return pd.concat(all_data) if all_data else pd.DataFrame()

def generate_mock_data(current_S):
    b1 = ["QA2K"]*150
    v1 = np.random.uniform(3.560,3.563,150)
    b2 = ["QBNJ"]*200
    v2 = np.random.uniform(3.552,3.555,200)
    df = pd.DataFrame({"Batch":b1+b2, "OCV":np.concatenate([v1,v2])})
    uni = df["Batch"].unique()
    if len(uni)>=2:
        a,b = uni[0],uni[1]
        maxv = df[df.Batch==a]["OCV"].max()
        add = pd.DataFrame({"Batch":[b]*current_S, "OCV":[maxv-0.0005]*current_S})
        df = pd.concat([df,add], ignore_index=True)
    return df

def sliding_max_min(arr, window):
    n = len(arr)
    max_vals = np.full(n - window + 1, np.nan)
    min_vals = np.full(n - window + 1, np.nan)
    for i in range(n - window + 1):
        seg = arr[i:i+window]
        max_vals[i] = seg.max()
        min_vals[i] = seg.min()
    return max_vals, min_vals

def perform_pairing_optimized(df, S, P, base_target, win_len, threshold, enable_s2, progress_bar):
    res_list = []
    meta_list = []
    cross_pack_info = []
    df = df.reset_index(drop=True).copy()
    df["Cell_ID"] = df.index
    used_ids = set()
    batches = df["Batch"].unique()
    total_batch = len(batches)
    batch_idx = 0

    for b in batches:
        batch_idx += 1
        progress_bar.progress(20 + int((batch_idx / total_batch) * 40), text=f"Stage1 處理批次 {b} ({batch_idx}/{total_batch})")
        full_sub = df[df["Batch"] == b].sort_values("OCV").reset_index(drop=True)
        remain_sub = full_sub.copy()
        while len(remain_sub) >= win_len:
            ocv_arr = remain_sub["OCV"].values
            max_win, min_win = sliding_max_min(ocv_arr, win_len)
            delta_arr = max_win - min_win
            valid_idx = np.where(delta_arr <= threshold)[0]
            if len(valid_idx) == 0:
                break
            best_pos = valid_idx[np.argmin(delta_arr[valid_idx])]
            best_window = remain_sub.iloc[best_pos : best_pos + win_len]
            cid_set = set(best_window["Cell_ID"])
            used_ids.update(cid_set)
            gid = f"P{len(res_list)+1}-{b}"
            pack_df = best_window.copy()
            pack_df["GroupID"] = gid
            pack_df["Method"] = "Stage 1"
            pack_df["Partner_Info"] = "單一批次內配組"
            res_list.append(pack_df)
            meta_list.append({
                "GroupID": gid,
                "Stage": "Stage 1",
                "Group_Min_OC(V)": round(min_win[best_pos],6),
                "Group_Max_OC(V)": round(max_win[best_pos],6),
                "Group_Delta(mV)": round((max_win[best_pos]-min_win[best_pos])*1000,2),
                "Cell_Count": win_len,
                "Batch_A": b,
                "Batch_B": None,
                "Cell_A": win_len,
                "Cell_B": 0
            })
            remain_sub = remain_sub[~remain_sub["Cell_ID"].isin(cid_set)].reset_index(drop=True)

    remain_all = df[~df["Cell_ID"].isin(used_ids)].reset_index(drop=True)
    progress_bar.progress(65, text="Stage1完成，進入Stage2混批組配")

    if enable_s2 and len(remain_all) >= win_len:
        batch_list = list(remain_all["Batch"].unique())
        comb_all = list(combinations(batch_list, 2))
        comb_total = len(comb_all)
        comb_idx = 0
        cache = {}
        for bb in batch_list:
            cache[bb] = remain_all[remain_all["Batch"] == bb].sort_values("OCV").reset_index(drop=True)
        for b1, b2 in comb_all:
            comb_idx += 1
            progress_bar.progress(65 + int((comb_idx / comb_total) * 20), text=f"Stage2 組合 {comb_idx}/{comb_total}: {b1}+{b2}")
            df1 = cache[b1]
            df2 = cache[b2]
            for split in range(1, P):
                p1 = split
                p2 = P - split
                take1 = S * p1
                take2 = S * p2
                if len(df1) < take1 or len(df2) < take2:
                    continue
                max1, min1 = sliding_max_min(df1["OCV"].values, take1)
                valid1 = np.where((max1 - min1) <= threshold)[0]
                if len(valid1) == 0:
                    continue
                max2, min2 = sliding_max_min(df2["OCV"].values, take2)
                valid2 = np.where((max2 - min2) <= threshold)[0]
                if len(valid2) == 0:
                    continue
                pair_candidates = []
                for i in valid1:
                    s1 = df1.iloc[i:i+take1]
                    for j in valid2:
                        s2 = df2.iloc[j:j+take2]
                        mix_all = pd.concat([s1, s2])
                        d_mix_max = mix_all["OCV"].max()
                        d_mix_min = mix_all["OCV"].min()
                        d_delta = d_mix_max - d_mix_min
                        if d_delta <= threshold:
                            pair_candidates.append((d_delta, i, j, s1, s2))
                if len(pair_candidates) == 0:
                    continue
                pair_candidates.sort()
                best_delta, i_best, j_best, s1_best, s2_best = pair_candidates[0]
                c1_set = set(s1_best["Cell_ID"])
                c2_set = set(s2_best["Cell_ID"])
                if c1_set & used_ids or c2_set & used_ids:
                    continue
                used_ids.update(c1_set)
                used_ids.update(c2_set)
                gid = f"S2-{b1}x{b2}-Sn{p1}"
                mix_df = pd.concat([s1_best, s2_best]).copy()
                mix_df["GroupID"] = gid
                mix_df["Method"] = "Stage 2"
                mix_df["Partner_Info"] = f"{b1}({p1}串) + {b2}({p2}串)"
                res_list.append(mix_df)
                meta_list.append({
                    "GroupID": gid,
                    "Stage": "Stage 2",
                    "Group_Min_OC(V)": round(d_mix_min,6),
                    "Group_Max_OC(V)": round(d_mix_max,6),
                    "Group_Delta(mV)": round(best_delta*1000,2),
                    "Cell_Count": take1 + take2,
                    "Batch_A": b1,
                    "Batch_B": b2,
                    "Cell_A": take1,
                    "Cell_B": take2
                })
                cross_pack_info.append({
                    "gid": gid,
                    "b1": b1,
                    "b2": b2,
                    "c1": take1,
                    "c2": take2,
                    "total": take1+take2
                })
                cache[b1] = df1[~df1["Cell_ID"].isin(c1_set)].reset_index(drop=True)
                cache[b2] = df2[~df2["Cell_ID"].isin(c2_set)].reset_index(drop=True)
    progress_bar.progress(85, text="分組完畢，生成統計表")
    final_group = pd.concat(res_list, ignore_index=True) if res_list else pd.DataFrame()
    unmatch = df[~df["Cell_ID"].isin(used_ids)].reset_index(drop=True)
    meta_df = pd.DataFrame(meta_list) if meta_list else pd.DataFrame()
    return final_group, unmatch, meta_df, used_ids, cross_pack_info

def generate_excel_report(raw_df, final_res, batch_report, group_summary, S, P, base_target, win_len, base_mv, enable_buffer, buffer_cell, tc, mt_unique, tp, yr):
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    font = "Segoe UI"
    h_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    z_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    a_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    thin = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
    w_font = Font(color="FFFFFFFF", bold=True, name=font)
    b_font = Font(bold=True, name=font)
    ws1 = wb.active
    ws1.title = "生產稽核總覽"
    ws1.merge_cells("A1:L1")
    ws1.cell(1,1,"🔋 電池配組精密稽核生產報告").font = Font(size=18,bold=True,name=font)
    table = [
        ["運行架構",f"{S}S{P}P","單PACK標準電芯",base_target],
        ["Buffer開啟","是" if enable_buffer else "否","Buffer數",buffer_cell],
        ["實際單PACK長度",win_len,"電壓門檻(mV)",base_mv],
        ["規格上限","3.00mV","總原始電芯",tc],
        ["成功配對唯一電芯",mt_unique,"真實總PACK組數",tp],
        ["良率",f"{yr:.1f}%","",""]
    ]
    for r,row in enumerate(table,3):
        for c,val in enumerate(row,1):
            cell = ws1.cell(r,c,val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c in [1,3]:
                cell.fill = a_fill
    ws1.cell(10,1,"📊 批次配組匯總").font = Font(size=14,bold=True,name=font)
    batch_header = ["批號","原始總數","Stage1獨立完整PACK(整數)","Stage1配組電芯數","混批組配等效Pack(權重)","混批配組電芯數","總配對(唯一)","未配對電芯","單批次良率"]
    for c,h in enumerate(batch_header,1):
        cell = ws1.cell(11,c,h)
        cell.fill = h_fill
        cell.font = w_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    for r_idx,b_data in enumerate(batch_report,12):
        row_vals = [
            b_data["批號(Batch)"],
            b_data["原始總數"],
            round(b_data["Stage1完整PACK"],4),
            b_data["Stage1電芯數"],
            round(b_data["混批組配等效Pack"],4),
            b_data["混批電芯數"],
            b_data["總配對(唯一)"],
            b_data["未配對"],
            f"{b_data['單批次良率']:.1f}%"
        ]
        for c,val in enumerate(row_vals,1):
            cell = ws1.cell(r_idx,c,val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r_idx % 2 == 0:
                cell.fill = z_fill
    if not group_summary.empty:
        ws_group = wb.create_sheet("PACK組電壓彙總")
        g_header = ["GroupID","配組階段","最小OC(V)","最大OC(V)","壓差(mV)","電芯數量","批次A","批次B","A貢獻電芯","B貢獻電芯"]
        for c,h in enumerate(g_header,1):
            cell = ws_group.cell(1,c,h)
            cell.fill = h_fill
            cell.font = w_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
        row_num = 2
        for _,row in group_summary.iterrows():
            vals = [
                row["GroupID"],
                row["Stage"],
                round(row["Group_Min_OC(V)"],6),
                round(row["Group_Max_OC(V)"],6),
                round(row["Group_Delta(mV)"],2),
                row["Cell_Count"],
                row["Batch_A"] if pd.notna(row["Batch_A"]) else "-",
                row["Batch_B"] if pd.notna(row["Batch_B"]) else "-",
                row["Cell_A"],
                row["Cell_B"]
            ]
            for col,val in enumerate(vals,1):
                cell = ws_group.cell(row_num, col, val)
                cell.border = thin
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if row_num % 2 == 0:
                    cell.fill = z_fill
            row_num += 1
    if not final_res.empty:
        ws_detail = wb.create_sheet("配組明細")
        d_header = ["Cell_ID","Batch","OC(V)","GroupID","配組階段","同組搭配批次"]
        for c,h in enumerate(d_header,1):
            cell = ws_detail.cell(1,c,h)
            cell.fill = h_fill
            cell.font = w_font
            cell.border = thin
        r = 2
        for _,row in final_res.iterrows():
            ws_detail.cell(r,1,row["Cell_ID"])
            ws_detail.cell(r,2,row["Batch"])
            ocv_cell = ws_detail.cell(r,3,row["OCV"])
            ocv_cell.number_format = "0.00000"
            ws_detail.cell(r,4,row["GroupID"])
            ws_detail.cell(r,5,row["Method"])
            ws_detail.cell(r,6,row["Partner_Info"])
            for col in range(1,7):
                cell = ws_detail.cell(r,col)
                cell.border = thin
                if r % 2 == 0:
                    cell.fill = z_fill
            r += 1
    ws_raw = wb.create_sheet("原始數據")
    ws_raw.cell(1,1,"Batch").fill = h_fill
    ws_raw.cell(1,1).font = w_font
    ws_raw.cell(1,2,"OC(V)").fill = h_fill
    ws_raw.cell(1,2).font = w_font
    r = 2
    for _,row in raw_df.iterrows():
        ws_raw.cell(r,1,row["Batch"])
        ocv_c = ws_raw.cell(r,2,row["OCV"])
        ocv_c.number_format = "0.00000"
        for c in [1,2]:
            cell = ws_raw.cell(r,c)
            cell.border = thin
            if r % 2 == 0:
                cell.fill = z_fill
        r += 1
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max_len + 4
    buf.seek(0)
    wb.save(buf)
    return buf.getvalue()

upload_files = st.file_uploader("📂 上傳Excel / CSV", accept_multiple_files=True, key=st.session_state.uploader_key)
df_upload = load_and_fix_data(upload_files)
st.session_state.upload_raw_data = df_upload

if not st.session_state.mock_data_df.empty:
    raw_df = st.session_state.mock_data_df
else:
    raw_df = st.session_state.upload_raw_data

col_btn1, col_btn2 = st.columns([1,1])
with col_btn1:
    mock_btn = st.button("🧪 載入模擬測試數據", key="mock_data_btn", use_container_width=True)
    if mock_btn:
        mock_data = generate_mock_data(S)
        st.session_state.mock_data_df = mock_data
        clear_analysis_keys = [
            "final_group_df", "remain_unmatch", "group_summary_df",
            "batch_report_list", "fig_main", "fig_scatter",
            "total_cells", "matched_unique", "unmatch_total",
            "total_pack_num", "yield_rate"
        ]
        df_keys = ["final_group_df", "remain_unmatch", "group_summary_df"]
        for k in clear_analysis_keys:
            if k in df_keys:
                st.session_state[k] = pd.DataFrame()
            else:
                st.session_state[k] = None
        st.success("✅ 已載入模擬批次數據，可執行配組")
        st.rerun()
with col_btn2:
    run_btn = st.button("🚀 執行配組分析", key="run_analysis_btn", use_container_width=True)

st.markdown("---")
if raw_df.empty:
    st.warning("⚠️ 無數據，請上傳檔案或點擊「載入模擬測試數據」按鈕")

if run_btn:
    if raw_df.empty:
        st.error("⚠️ 無數據，請上傳檔案或載入模擬數據")
        st.stop()
    progress_bar = st.progress(0, text="演算配組中...")
    with st.spinner("正在分組計算，請稍候..."):
        total_cells = len(raw_df)
        calc_df = raw_df.reset_index(drop=True).copy()
        calc_df["Cell_ID"] = calc_df.index
        final_group_df, remain_unmatch, group_summary_df, used_set, cross_pack_list = perform_pairing_optimized(
            calc_df, S, P, base_target, win_len, base_threshold, enable_s2, progress_bar
        )
        progress_bar.progress(90, text="生成統計圖表與報告數據")
        matched_unique = len(used_set)
        unmatch_total = len(remain_unmatch)
        total_pack_num = final_group_df["GroupID"].nunique() if not final_group_df.empty else 0
        yield_rate = (matched_unique / total_cells)*100 if total_cells>0 else 0

        st.session_state.total_cells = total_cells
        st.session_state.matched_unique = matched_unique
        st.session_state.unmatch_total = unmatch_total
        st.session_state.total_pack_num = total_pack_num
        st.session_state.yield_rate = yield_rate
        st.session_state.final_group_df = final_group_df
        st.session_state.remain_unmatch = remain_unmatch
        st.session_state.group_summary_df = group_summary_df

        batch_report_list = []
        chart_data = []
        cross_cell_map = {}
        for item in cross_pack_list:
            b1 = item["b1"]
            b2 = item["b2"]
            c1 = item["c1"]
            c2 = item["c2"]
            if b1 not in cross_cell_map:
                cross_cell_map[b1] = 0
            cross_cell_map[b1] += c1
            if b2 not in cross_cell_map:
                cross_cell_map[b2] = 0
            cross_cell_map[b2] += c2

        for b in raw_df["Batch"].unique():
            b_all = len(raw_df[raw_df["Batch"]==b])
            batch_calc = calc_df[calc_df["Batch"] == b]
            batch_used = set(used_set) & set(batch_calc["Cell_ID"])
            match_unique = len(batch_used)
            un = b_all - match_unique
            s1_df = final_group_df[final_group_df["Method"]=="Stage 1"]
            s1_full = s1_df[s1_df["Batch"]==b]["GroupID"].nunique()
            s1_cell = s1_full * win_len
            s2_cell = cross_cell_map.get(b, 0)
            y = (match_unique/b_all)*100 if b_all>0 else 0.0
            batch_report_list.append({
                "批號(Batch)":b,
                "原始總數":b_all,
                "Stage1完整PACK": s1_full,
                "Stage1電芯數": s1_cell,
                "混批組配等效Pack": round(s2_cell / win_len,4),
                "混批電芯數": s2_cell,
                "總配對(唯一)":match_unique,
                "未配對":un,
                "單批次良率":y
            })
            # 堆疊圖三類數據（電芯數）
            chart_data.append({"批次":b,"電芯數":s1_cell,"分類":"Stage1獨立PACK"})
            chart_data.append({"批次":b,"電芯數":s2_cell,"分類":"混批組配等效Pack"})
            chart_data.append({"批次":b,"電芯數":un,"分類":"無法配組電芯"})
        st.session_state.batch_report_list = batch_report_list
        # 單張堆疊長條圖，X軸統一電芯數，開啟數值標籤
        df_chart = pd.DataFrame(chart_data)
        fig_main = px.bar(df_chart, x="電芯數", y="批次", color="分類", orientation="h",
                          barmode="stack", text="電芯數", text_auto=True,
                          color_discrete_map={
                              "Stage1獨立PACK":"#2ecc71",
                              "混批組配等效Pack":"#f39c12",
                              "無法配組電芯":"#e74c3c"
                          })
        fig_main.update_layout(title="各批次電芯分佈堆疊長條圖(Stage1/混批/剩餘未配電芯)", height=400, plot_bgcolor="white")
        fig_main.update_traces(texttemplate="%{text}", textposition="inside")
        st.session_state.fig_main = fig_main
        # 全域電壓散佈圖不變
        sc_all = pd.concat([final_group_df, remain_unmatch.assign(GroupID="Unmatched",Method="Unmatched",Partner_Info="-")], ignore_index=True)
        sc_all["mV"] = sc_all["OCV"]*1000
        sc_all["Seq"] = list(range(len(sc_all)))
        fig_scatter = px.scatter(sc_all, x="Seq", y="mV", color="GroupID", color_discrete_map=FIX_COLOR_MAP, hover_data=["Batch","Method","Partner_Info"], labels={"Seq":"電芯序號","mV":"電壓(mV)"})
        fig_scatter.update_traces(marker=dict(size=6,opacity=0.85))
        fig_scatter.update_layout(height=480, plot_bgcolor="white")
        st.session_state.fig_scatter = fig_scatter
        progress_bar.progress(100, text="演算完成！")
        st.success("✅ 配組演算完畢，下方顯示結果")

# 頁面渲染
if not raw_df.empty and st.session_state.final_group_df is not None and not st.session_state.final_group_df.empty:
    tc = st.session_state.total_cells
    mt_unique = st.session_state.matched_unique
    tp = st.session_state.total_pack_num
    yr = st.session_state.yield_rate
    gs = st.session_state.group_summary_df
    br = st.session_state.batch_report_list
    fig_main = st.session_state.fig_main
    fs = st.session_state.fig_scatter

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("總電芯", f"{tc} 顆")
    c2.metric("配對唯一電芯", f"{mt_unique} 顆")
    c3.metric("真實總PACK組數", f"{tp} 組")
    c4.metric("整體良率", f"{yr:.1f}%")
    st.markdown("---")
    st.subheader("PACK電壓一覽")
    st.dataframe(gs, use_container_width=True)
    st.markdown("---")
    st.subheader("各批次電芯分佈堆疊長條圖")
    st.plotly_chart(fig_main, use_container_width=True)
    st.markdown("""
    > 圖表說明：
    > 1. X軸統一為「電芯顆數」，三類數據堆疊展示；
    > 2. 綠色：Stage1同批獨立配組總電芯數；橙色：混批組配總貢獻電芯數；紅色：無法配組剩餘電芯；
    > 3. 每一段長條內自動標註該段對應電芯數量。
    """)
    st.markdown("---")
    st.subheader("批次統計表")
    st.dataframe(pd.DataFrame(br), use_container_width=True)
    st.markdown("---")
    st.subheader("全域電壓散佈圖")
    st.plotly_chart(fs, use_container_width=True)
    st.markdown("---")
    st.subheader("報告下載")
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_bytes = generate_excel_report(raw_df, st.session_state.final_group_df, br, gs, S, P, base_target, win_len, base_mv, enable_buffer, buffer_cell, tc, mt_unique, tp, yr)
    st.download_button("📥 下載Excel報告", excel_bytes, f"配組報告_{ts}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.success("✅ 分析完成，僅支援Excel下載")